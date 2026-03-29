from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from config import KERF_MM, OVERMEASURE_MM
from logic import Cabinet, Element, MaterialSummary


def _style_header(ws, row):
    for cell in ws[row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="222222")
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autosize(ws):
    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col[0].column_letter].width = max(12, max_len + 2)


def export_excel(
    cabinet: Cabinet,
    elements: list[Element],
    summary: MaterialSummary,
    file_path: str,
    carcass_material: str,
    back_material: str,
):
    wb = Workbook()

    # =========================
    # PRIREZOVALNA LISTA
    # =========================
    ws = wb.active
    ws.title = "PRIREZOVALNA LISTA"

    ws.merge_cells("A1:M1")
    ws["A1"] = f"PRIREZOVALNA LISTA {cabinet.width} x {cabinet.height} x {cabinet.depth}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.append([
        "poz.", "št. kos", "Naziv elementa", "Material",
        "Neobdelan L", "W", "Deb.",
        "m²", "m³",
        "Obdelan L", "W", "Deb.",
        "Opombe"
    ])
    _style_header(ws, 2)

    pos = 1

    for el in elements:

        # NAZIVI
        if el.name.lower() == "stranica":
            naziv = "L in D stranica"
        elif el.name.lower() == "strop":
            naziv = "strop in dno"
        elif el.name.lower() == "dno":
            continue  # že vključeno zgoraj
        elif el.name.lower() == "polica":
            naziv = "polica"
        elif "hrbt" in el.name.lower():
            naziv = "hrbtišče"
        else:
            naziv = el.name

        material = back_material if "hrbt" in el.name.lower() else carcass_material

        raw_l = el.width + OVERMEASURE_MM
        raw_w = el.height + OVERMEASURE_MM

        area = (el.width * el.height * el.quantity) / 1_000_000
        volume = (el.width * el.height * el.thickness * el.quantity) / 1_000_000_000

        ws.append([
            pos,
            el.quantity,
            naziv,
            material,
            raw_l,
            raw_w,
            el.thickness,
            round(area, 2),
            round(volume, 5),
            el.width,
            el.height,
            el.thickness,
            "žaganje z nadmerami"
        ])

        pos += 1

    ws.append([])
    ws.append([f"Kerf (rez žage): {KERF_MM} mm"])

    _autosize(ws)

    # =========================
    # MATERIALNA LISTA
    # =========================
    ws2 = wb.create_sheet("MATERIALNA LISTA")

    ws2.merge_cells("A1:E1")
    ws2["A1"] = f"MATERIALNA LISTA {cabinet.width} x {cabinet.height} x {cabinet.depth}"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2["A1"].alignment = Alignment(horizontal="center")

    ws2.append([
        "Material", "Dimenzija mm", "Enota", "Količina", "Opombe"
    ])
    _style_header(ws2, 2)

    ws2.append([
        "Oplemenitena iverna plošča", "19", "m²",
        round(summary.surface_m2, 2), "Kronospan"
    ])

    ws2.append([
        "Vlaknena plošča (HDF)", "4", "m²",
        round(summary.surface_m2 * 0.3, 2), ""
    ])

    ws2.append([
        "Les (smreka)", "-", "m³",
        round(summary.volume_m3, 5), ""
    ])

    ws2.append([
        "Robni trak", "-", "tm",
        round(summary.edge_banding_m, 2), "Hranipex"
    ])

    ws2.append([
        "Mozniki Ø8x32", "-", "kos",
        summary.dowels_count, ""
    ])

    ws2.append([
        "Lepilo PVAc", "-", "kg",
        round(summary.dowels_count * 0.005, 3), ""
    ])

    ws2.append([
        "Vijaki", "-", "kos",
        summary.dowels_count // 2, ""
    ])

    _autosize(ws2)

    wb.save(file_path)